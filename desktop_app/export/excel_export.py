from copy import deepcopy

from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from scheduler.api.utils.holidays import get_holidays_for_month



RED_FILL = PatternFill("solid", fgColor="FF6666")
GREEN_FILL = PatternFill("solid", fgColor="99CC66")
HEADER_FILL = PatternFill("solid", fgColor="E6E6E6")

EMPLOYEE_ROW_FILLS = [
    PatternFill("solid", fgColor="FFD966"),
    PatternFill("solid", fgColor="9BC2E6"),
    PatternFill("solid", fgColor="A9D18E"),
    PatternFill("solid", fgColor="F4B084"),
    PatternFill("solid", fgColor="C5E0B4"),
]

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


COUNT_AS_WORKED = {"Д", "В", "Н", "А", "О", "Б"}


def export_schedule_to_excel(
    filename: str,
    company: str,
    department: str,
    city: str,
    month_name: str,
    month: int,
    year: int,
    employees: list[dict],
    days: list[int],
    schedule: dict,
):
    schedule = deepcopy(schedule)

    wb = Workbook()
    ws = wb.active
    ws.title = "График"

    holidays = set(get_holidays_for_month(year, month))

    FIRST_DAY_COL = 4
    LAST_DAY_COL = FIRST_DAY_COL + len(days) - 1
    CARD_COL = LAST_DAY_COL + 1

    # ===== HEADER =====
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=10)
    ws.cell(3, 3, company).font = Font(size=14, bold=True)
    ws.cell(3, 3).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=3, start_column=11, end_row=3, end_column=22)
    ws.cell(3, 11, department).font = Font(size=16, bold=True)
    ws.cell(3, 11).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=3, start_column=23, end_row=3, end_column=CARD_COL)
    ws.cell(3, 23, city).font = Font(size=14, bold=True)
    ws.cell(3, 23).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=CARD_COL)
    ws.cell(4, 1, f"{month_name} {year} г.").font = Font(size=12, bold=True)
    ws.cell(4, 1).alignment = Alignment(horizontal="center")


    headers = ["№", "Бр.", "Име, Презиме, Фамилия"]
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col, value=text)
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")


    for idx, day in enumerate(days):
        col = FIRST_DAY_COL + idx
        c = ws.cell(row=5, column=col, value=day)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

        weekday = date(year, month, day).weekday()
        c.fill = RED_FILL if (weekday >= 5 or day in holidays) else GREEN_FILL

    ws.cell(row=5, column=CARD_COL, value="Служебен №").fill = HEADER_FILL
    ws.cell(row=5, column=CARD_COL).border = THIN_BORDER
    ws.cell(row=5, column=CARD_COL).alignment = Alignment(horizontal="center")

    # ===== BODY =====
    for idx, emp in enumerate(employees):
        row = 6 + idx

        emp_id = str(emp["id"])
        name = emp["full_name"]
        card = emp.get("card_number", "")

        emp_days = schedule.get(emp_id, {})

        worked_days = sum(
            1 for d in days if emp_days.get(d, "") in COUNT_AS_WORKED
        )

        fill = EMPLOYEE_ROW_FILLS[idx % len(EMPLOYEE_ROW_FILLS)]

        ws.cell(row=row, column=1, value=idx + 1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=worked_days).border = THIN_BORDER

        name_cell = ws.cell(row=row, column=3, value=name)
        name_cell.border = THIN_BORDER
        name_cell.fill = fill

        for i, day in enumerate(days):
            col = FIRST_DAY_COL + i
            val = emp_days.get(day, "")
            c = ws.cell(row=row, column=col, value=val)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=CARD_COL, value=card).border = THIN_BORDER

    # ===== LEGEND =====
    legend_row = row + 3
    ws.cell(legend_row, 3, "ЛЕГЕНДА:").font = Font(bold=True)

    legend = [
        "Д – ДНЕВНА (08:00–16:00)",
        "В – ВТОРА (16:00–24:00)",
        "Н – НОЩНА (00:00–08:00)",
        "А – АДМИНИСТРАЦИЯ",
        "О – ОТПУСК",
        "Б – БОЛНИЧЕН",
        "ПРАЗНО – ПОЧИВЕН ДЕН",
    ]

    for i, text in enumerate(legend, start=1):
        ws.cell(legend_row + i, 3, text)

    # ===== WIDTHS =====
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 34

    for col in range(FIRST_DAY_COL, CARD_COL):
        ws.column_dimensions[get_column_letter(col)].width = 4

    ws.column_dimensions[get_column_letter(CARD_COL)].width = 16

    wb.save(filename)
